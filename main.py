import pandas as pd
import numpy as np
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from sklearn.preprocessing import StandardScaler
from datetime import datetime, timedelta
import random
import socket
import geocoder
import openpyxl
from openpyxl.styles import PatternFill

# Get system IP address and location
def get_system_info():
    ip_address = socket.gethostbyname(socket.gethostname())
    g = geocoder.ip(ip_address)
    location = g.city if g.city else "Unknown City"
    region = g.country if g.country else "Unknown Country"
    return ip_address, location, region

# Simulate historical cyber attack data based on system's IP addresses
def generate_data(num_records=1000):
    np.random.seed(42)
    random.seed(42)
    
    ip_address, location, region = get_system_info()
    attack_types = ['DDoS', 'Phishing', 'Malware', 'Ransomware','Cloud-based-attacks','Data-center-attacks','Passwaord attacks','web attacks','Trojan horses']
    ports = [80, 443, 21, 22, 25, 8080]
    protocols = ['TCP', 'UDP', 'ICMP','HTTP','FTP','SMTP','IP']
    
    data = []
    
    for _ in range(num_records):
        attack_type = random.choice(attack_types)
        port = random.choice(ports)
        timestamp = datetime.now() - timedelta(days=random.randint(0, 365))
        attack_duration = random.randint(1, 3600)  # in seconds
        impact = random.randint(1, 10)
        success_level = random.randint(1, 10)
        protocol = random.choice(protocols)
        
        severity_score = impact * success_level  # Simplified severity score calculation
        
        data.append([severity_score, ip_address, attack_type, port, timestamp, attack_duration, region, impact, success_level, protocol])
    
    columns = ['SeverityScore', 'IPAddress', 'AttackType', 'Port', 'Timestamp', 'AttackDuration', 'Region', 'Impact', 'SuccessLevel', 'Protocol']
    df = pd.DataFrame(data, columns=columns)
    
    return df

# Generate and save historical data
historical_data = generate_data()
historical_data.to_csv('historical_cyber_attacks.csv', index=False)

# Load historical data
data = pd.read_csv('historical_cyber_attacks.csv')

# Preprocess data
data['Timestamp'] = pd.to_datetime(data['Timestamp'])
data_encoded = pd.get_dummies(data, columns=['AttackType', 'Region', 'Protocol'])

# Save original columns before converting to NumPy array
original_columns = data_encoded.drop(columns=['SeverityScore', 'IPAddress', 'Timestamp']).columns

# Split data into features and labels
X = data_encoded.drop(columns=['SeverityScore', 'IPAddress', 'Timestamp'])
y = data_encoded['SeverityScore']

# Normalize data
scaler = StandardScaler()
X = scaler.fit_transform(X)

# Split into training and testing sets
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

# Train a Random Forest model
model = RandomForestClassifier(n_estimators=100, random_state=42)
model.fit(X_train, y_train)

# Make predictions and evaluate
y_pred = model.predict(X_test)
print("Model Accuracy:", accuracy_score(y_test, y_pred))

# Predict future threats (simulated data for the next month)
def predict_future_threats(model, scaler, original_columns, num_days=30):
    future_data = generate_data(num_records=num_days)
    future_data['Timestamp'] = pd.to_datetime(future_data['Timestamp'])
    future_data_encoded = pd.get_dummies(future_data, columns=['AttackType', 'Region', 'Protocol'])
    
    # Ensure future data has same columns as training data
    for col in original_columns:
        if col not in future_data_encoded.columns:
            future_data_encoded[col] = 0
    future_data_encoded = future_data_encoded[original_columns]
    
    future_X = scaler.transform(future_data_encoded)
    future_data['PredictedSeverityScore'] = model.predict(future_X)
    
    # Predict future attack type based on the most frequent type predicted
    future_data['PredictedAttackType'] = future_data_encoded[['AttackType_DDoS', 'AttackType_Phishing', 'AttackType_Malware', 'AttackType_Ransomware']].idxmax(axis=1).str.replace('AttackType_', '')

    return future_data

future_threats = predict_future_threats(model, scaler, original_columns)

# Save future threats predictions to CSV
future_threats.to_csv('predicted_future_threats.csv', index=False)

# Generate mitigation measures (simplified example)
def suggest_mitigation_measures(threats):
    mitigation_measures = []
    for _, row in threats.iterrows():
        if row['PredictedSeverityScore'] > 70:
            measure = "Immediate action required: Isolate the affected systems and begin incident response procedures."
        elif row['PredictedSeverityScore'] > 40:
            measure = "High priority: Monitor the systems closely and prepare for potential incident response."
        else:
            measure = "Low priority: Regular monitoring and standard security practices."
        mitigation_measures.append(measure)
    
    threats['MitigationMeasures'] = mitigation_measures
    return threats

future_threats_with_measures = suggest_mitigation_measures(future_threats)

# Save final report to CSV
future_threats_with_measures.to_csv('cyber_threats_report.csv', index=False)

# Display future threat message with attack type
future_attack_type = future_threats['PredictedAttackType'].mode()[0]


# Save reports with formatting
def save_report_with_formatting(df, filename):
    df.to_excel(filename, index=False)
    
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # Define fill colors for conditional formatting
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=23):  # Extend the color to the first 23 columns
        severity_score = row[0].value  # Assuming SeverityScore is the first column
        if severity_score is not None:
            if severity_score < 40:
                fill = green_fill
            elif severity_score < 70:
                fill = yellow_fill
            else:
                fill = red_fill
            for cell in row:
                cell.fill = fill  # Apply the fill to the first 23 columns of the row
    
    wb.save(filename)

save_report_with_formatting(future_threats_with_measures, 'predicted_future_threats_report.xlsx')
save_report_with_formatting(future_threats_with_measures, 'cyber_threats_report_file.xlsx')
save_report_with_formatting(future_threats_with_measures, 'historical_cyber_attacks.xlsx')
print ("Reports Generated Successfully ")

